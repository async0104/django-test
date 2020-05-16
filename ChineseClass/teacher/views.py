from django.shortcuts import render
from django.views.decorators.http import require_http_methods
from django.core import serializers
from django.http import JsonResponse
from django.forms.models import model_to_dict
from django.contrib.auth.hashers import make_password, check_password
import json
import xlrd
from openpyxl import load_workbook
import openpyxl
from . models import Teacher,Question
from student.models import Student,Student_Class
from .models import Classes
from control.models import Class_permission
from control.models import School
from .models import Homework,Question
from teacher.models import Homework_score, Question_bank, Sub_answer
# Create your views here.

#教师登录
@require_http_methods(["POST"])
def login(request):
    response = {}
    result = json.loads(request.body.decode())
    try:
        teacher = Teacher.objects.get(username = result.get('username'))
        if(check_password(result.get('password'),teacher.password)):
            response['msg'] = 'success'
        else:
            response['msg'] = 'passwordfailure'
    except  Exception as e:
        response['msg'] = 'usernamefailure'
    return JsonResponse(response)

#教师注册
@require_http_methods(["POST"])
def register(request):
    response = {}
    result = json.loads(request.body.decode())
    try:
        school = School.objects.get(name = result.get('school'))
    except  Exception as e:
        response['msg'] = 'school_failure'
        return JsonResponse(response)
    try:
        teacher = Teacher (
            username = result.get('username'),
            password = make_password(result.get('password')),
            realname = result.get('realname'),
            sex = result.get('sex'),
            age = result.get('age'),
            phone = result.get('phone'),
            subject = result.get('subject'),
            userIcon = result.get('userIcon'),
            school = school
        )
        teacher.save()
        response['msg'] = 'success'
    except  Exception as e:
        response['msg'] = 'failure'
    return JsonResponse(response)

#教师查看个人信息
@require_http_methods(["GET"])
def show_information(request):
    response = {}
    try:
        teacher = Teacher.objects.get(username = request.GET.get('username'))
    except Exception as e:
        response['msg'] = 'failure' #教师不存在
        return  JsonResponse(response)
    response['username'] = teacher.username
    response['realname'] = teacher.realname
    response['sex'] = teacher.sex
    response['age'] = teacher.age
    response['phone'] = teacher.phone
    response['subject'] = teacher.subject
    response['userIcon'] = teacher.userIcon
    response['school'] = teacher.school.name
    return  JsonResponse(response)

#创建班级
@require_http_methods(["GET"])
def create_class(request):
    response = {}
    try:
        permission = Class_permission.objects.get(name = request.GET.get('permission'))
    except  Exception as e:
        response['msg'] = 'permission_failure'
        return JsonResponse(response)
    
    if  permission.use == True:
        response['msg'] = 'permission_used_failure'
        return JsonResponse(response)

    try:
        teacher1 = Teacher.objects.get(username = request.GET.get('username'))
    except  Exception as e:
        response['msg'] = 'teacher_failure'
        return JsonResponse(response)
    
    try:
        classes = Classes(
            name = request.GET.get('name'),
            description = request.GET.get('description'),
            permission = permission,
            teacher = teacher1
        )
        classes.save()
        Class_permission.objects.filter(id = permission.id).update(use = True)
        response['msg'] = 'success'
    except  Exception as e:
        response['msg'] = 'failure'
    return JsonResponse(response)

#教师查看自己创建的班级
@require_http_methods(["GET"])
def show_classes(request):
    response = {}
    try:
        teacher = Teacher.objects.get(username = request.GET.get('username'))
    except  Exception as e:
        response['msg'] = 'teacher_failure' #教师不存在
        return JsonResponse(response)
    names = []
    classes = Classes.objects.filter(teacher = teacher)
    num = Classes.objects.filter(teacher = teacher).count()
    for i in range(0, num):
        names.append(classes[i].name)
    response['list'] = names
    return JsonResponse(response)

#教师点击某个班级查看详细信息(根据班级名字)
@require_http_methods(["GET"])
def show_class_infor(request):
    response = {}
    try:
        classes = Classes.objects.get(name = request.GET.get('name'))
    except Exception as e:
        response['msg'] = 'class_failure' #班级不存在
        return JsonResponse(response)
    response['description'] = classes.description
    response['createdate'] = classes.create_date
    try:
        studentNames = []
        student_usernames = []
        student_class = Student_Class.objects.filter(classes = classes)
        for i in range(0, student_class.count()):
            studentNames.append(student_class[i].student.realname)
            student_usernames.append(student_class[i].student.username)
    except Exception as e:
        response['msg'] = str(e)
        return JsonResponse(response)
    names = []
    homework = Homework.objects.filter(classes = classes)
    num = Homework.objects.filter(classes = classes).count()
    for i in range(0, num):
        names.append(homework[i].name)
    response['student_realname'] = studentNames
    response['student_username'] = student_usernames
    response['homework'] = names
    response['msg'] = 'success'
    return JsonResponse(response)

#查看学生信息
@require_http_methods(["GET"])
def show_student_information(request):
    response = {}
    try:
        student = Student.objects.get(username = request.GET.get('username'))
    except Exception as e:
        response['msg'] = 'failure' #学生不存在
        return  JsonResponse(response)
    response['username'] = student.username
    response['realname'] = student.realname
    response['sex'] = student.sex
    response['age'] = student.age
    response['phone'] = student.phone
    response['grade'] = student.grade
    response['userIcon'] = student.userIcon
    return  JsonResponse(response)

#删除学生
@require_http_methods(["GET"])
def delete_student(request):
    response = {}
    try:
        student = Student.objects.get(username = request.GET.get('username'))
    except Exception as e:
        response['msg'] = 'student_failure' #学生不存在
        return  JsonResponse(response)
    
    try:
        student.delete()
        response['msg'] = 'success'
    except Exception as e:
        response['msg'] = 'failure' #删除失败
    return  JsonResponse(response)
    
#教师点击某个作业，查看学生得分情况

#布置作业(题号方式)
@require_http_methods(["POST"])
def assign_homework_by_number(request):
    response = {}
    result = json.loads(request.body.decode())
    locations = result.get('location')
    classes = Classes.objects.get(name = result.get('class'))
    try:
        homework = Homework(
            name = result.get('name'),
            classes = classes,
            close_time = result.get('close_time'),
        )
        homework.save()
    except  Exception as e:
        response['msg'] = 'name_failure' #作业名称已存在
        JsonResponse(response)
    try:
        for location in locations:
            question = Question.objects.get(location = location)
            homework.question.add(question)
    except  Exception as e:
        response['msg'] = 'question_failure'#题目不存在
        return JsonResponse(response)
    response['msg'] = 'success'
    return JsonResponse(response)

#教师选题
@require_http_methods(["GET"])
def show_questions(request):
    response = {}
    book = xlrd.open_workbook('C:\\excel\\question_bank.xlsx')
    sheet1 = book.sheets()[0]
    nrows = sheet1.nrows #题库表格总行数
    try:
        types = request.GET.get('type')
        subject = request.GET.get('subject')
        homework = []
        if(types == "single"): #单选
            if(subject == "Math"): #数学
                questions = Question.objects.filter(types = "单选", subject = "数学")
                num = Question.objects.filter(types = "单选", subject = "数学").count()
                for i in range(0, num):
                    for j in range(0, nrows): #遍历整个excel表
                        if(questions[i].location == j+1): #找到了对应的题目
                            d = {"order": j+1, "question": sheet1.cell(j,1).value, "choice1": sheet1.cell(j,2).value, 
                            "choice2": sheet1.cell(j,3).value, "choice3": sheet1.cell(j,4).value, "choice4": sheet1.cell(j,5).value
                            ,"answer": sheet1.cell(j,6).value, "description": sheet1.cell(j,7).value}    
                            homework.append(d)
            elif(subject == "Chinese"): #语文
                questions = Question.objects.filter(types = "单选", subject = "语文")
                num = Question.objects.filter(types = "单选", subject = "语文").count()
                for i in range(0, num):
                    for j in range(0, nrows): 
                        if(questions[i].location == j+1): 
                            d = {"order": j+1, "question": sheet1.cell(j,1).value, "choice1": sheet1.cell(j,2).value, 
                            "choice2": sheet1.cell(j,3).value, "choice3": sheet1.cell(j,4).value, "choice4": sheet1.cell(j,5).value
                            ,"answer": sheet1.cell(j,6).value, "description": sheet1.cell(j,7).value}    
                            homework.append(d)
            elif(subject == "English"): #英语
                questions = Question.objects.filter(types = "单选", subject = "英语")
                num = Question.objects.filter(types = "单选", subject = "英语").count()
                for i in range(0, num):
                    for j in range(0, nrows): 
                        if(questions[i].location == j+1): 
                            d = {"order": j+1, "question": sheet1.cell(j,1).value, "choice1": sheet1.cell(j,2).value, 
                            "choice2": sheet1.cell(j,3).value, "choice3": sheet1.cell(j,4).value, "choice4": sheet1.cell(j,5).value
                            ,"answer": sheet1.cell(j,6).value, "description": sheet1.cell(j,7).value}    
                            homework.append(d)
        elif(types == "mul"): #多选
            if(subject == "Math"): #数学
                questions = Question.objects.filter(types = "多选", subject = "数学")
                num = Question.objects.filter(types = "多选", subject = "数学").count()
                for i in range(0, num):
                    for j in range(0, nrows): #遍历整个excel表
                        if(questions[i].location == j+1): #找到了对应的题目
                            d = {"order": j+1, "question": sheet1.cell(j,1).value, "choice1": sheet1.cell(j,2).value, 
                            "choice2": sheet1.cell(j,3).value, "choice3": sheet1.cell(j,4).value, "choice4": sheet1.cell(j,5).value,
                            "choice5": sheet1.cell(j,6).value, "answer": sheet1.cell(j,7).value, "description": sheet1.cell(j,8).value}    
                            homework.append(d)
            elif(subject == "Chinese"): #语文
                questions = Question.objects.filter(types = "多选", subject = "语文")
                num = Question.objects.filter(types = "多选", subject = "语文").count()
                for i in range(0, num):
                    for j in range(0, nrows): 
                        if(questions[i].location == j+1): 
                            d = {"order": j+1, "question": sheet1.cell(j,1).value, "choice1": sheet1.cell(j,2).value, 
                            "choice2": sheet1.cell(j,3).value, "choice3": sheet1.cell(j,4).value, "choice4": sheet1.cell(j,5).value,
                            "choice5": sheet1.cell(j,6).value, "answer": sheet1.cell(j,7).value, "description": sheet1.cell(j,8).value}    
                            homework.append(d)
            elif(subject == "English"): #英语
                questions = Question.objects.filter(types = "多选", subject = "英语")
                num = Question.objects.filter(types = "多选", subject = "英语").count()
                for i in range(0, num):
                    for j in range(0, nrows): 
                        if(questions[i].location == j+1): 
                            d = {"order": j+1, "question": sheet1.cell(j,1).value, "choice1": sheet1.cell(j,2).value, 
                            "choice2": sheet1.cell(j,3).value, "choice3": sheet1.cell(j,4).value, "choice4": sheet1.cell(j,5).value,
                            "choice5": sheet1.cell(j,6).value, "answer": sheet1.cell(j,7).value, "description": sheet1.cell(j,8).value}    
                            homework.append(d)
        elif(types == "completion"): #填空
            if(subject == "Math"): #数学
                questions = Question.objects.filter(types = "填空", subject = "数学")
                num = Question.objects.filter(types = "填空", subject = "数学").count()
                for i in range(0, num):
                    for j in range(0, nrows): #遍历整个excel表
                        if(questions[i].location == j+1): #找到了对应的题目
                            d = {"order": j+1, "question": sheet1.cell(j,1).value, "answer": sheet1.cell(j,2).value, 
                            "description": sheet1.cell(j,3).value}    
                            homework.append(d)
            elif(subject == "Chinese"): #语文
                questions = Question.objects.filter(types = "填空", subject = "语文")
                num = Question.objects.filter(types = "填空", subject = "语文").count()
                for i in range(0, num):
                    for j in range(0, nrows): #遍历整个excel表
                        if(questions[i].location == j+1): #找到了对应的题目
                            d = {"order": j+1, "question": sheet1.cell(j,1).value, "answer": sheet1.cell(j,2).value, 
                            "description": sheet1.cell(j,3).value}    
                            homework.append(d)
            elif(subject == "English"): #英语
                questions = Question.objects.filter(types = "填空", subject = "英语")
                num = Question.objects.filter(types = "填空", subject = "英语").count()
                for i in range(0, num):
                    for j in range(0, nrows): #遍历整个excel表
                        if(questions[i].location == j+1): #找到了对应的题目
                            d = {"order": j+1, "question": sheet1.cell(j,1).value, "answer": sheet1.cell(j,2).value, 
                            "description": sheet1.cell(j,3).value}    
                            homework.append(d)
        elif(types == "judgement"): #判断
            if(subject == "Math"): #数学
                questions = Question.objects.filter(types = "判断", subject = "数学")
                num = Question.objects.filter(types = "判断", subject = "数学").count()
                for i in range(0, num):
                    for j in range(0, nrows): #遍历整个excel表
                        if(questions[i].location == j+1): #找到了对应的题目
                            d = {"order": j+1, "question": sheet1.cell(j,1).value, "answer": sheet1.cell(j,2).value, 
                            "description": sheet1.cell(j,3).value}   
                            homework.append(d)
            elif(subject == "Chinese"): #语文
                questions = Question.objects.filter(types = "判断", subject = "语文")
                num = Question.objects.filter(types = "判断", subject = "语文").count()
                for i in range(0, num):
                    for j in range(0, nrows): #遍历整个excel表
                        if(questions[i].location == j+1): #找到了对应的题目
                            d = {"order": j+1, "question": sheet1.cell(j,1).value, "answer": sheet1.cell(j,2).value, 
                            "description": sheet1.cell(j,3).value}   
                            homework.append(d)
            elif(subject == "English"): #英语
                questions = Question.objects.filter(types = "判断", subject = "英语")
                num = Question.objects.filter(types = "判断", subject = "英语").count()
                for i in range(0, num):
                    for j in range(0, nrows): #遍历整个excel表
                        if(questions[i].location == j+1): #找到了对应的题目
                            d = {"order": j+1, "question": sheet1.cell(j,1).value, "answer": sheet1.cell(j,2).value, 
                            "description": sheet1.cell(j,3).value}   
                            homework.append(d)
        elif(types == "subjective"): #主观
            if(subject == "Math"): #数学
                questions = Question.objects.filter(types = "主观", subject = "数学")
                num = Question.objects.filter(types = "主观", subject = "数学").count()
                for i in range(0, num):
                    for j in range(0, nrows): #遍历整个excel表
                        if(questions[i].location == j+1): #找到了对应的题目
                            d = {"order": j+1, "question": sheet1.cell(j,1).value, "description": sheet1.cell(j,2).value}   
                            homework.append(d)
            elif(subject == "Chinese"): #语文
                questions = Question.objects.filter(types = "主观", subject = "语文")
                num = Question.objects.filter(types = "主观", subject = "语文").count()
                for i in range(0, num):
                    for j in range(0, nrows): #遍历整个excel表
                        if(questions[i].location == j+1): #找到了对应的题目
                            d = {"order": j+1, "question": sheet1.cell(j,1).value, "description": sheet1.cell(j,2).value}   
                            homework.append(d)
            elif(subject == "English"): #英语
                questions = Question.objects.filter(types = "主观", subject = "英语")
                num = Question.objects.filter(types = "主观", subject = "英语").count()
                for i in range(0, num):
                    for j in range(0, nrows): #遍历整个excel表
                        if(questions[i].location == j+1): #找到了对应的题目
                            d = {"order": j+1, "question": sheet1.cell(j,1).value, "description": sheet1.cell(j,2).value}
                            homework.append(d)
        response['list'] = homework
        response['msg'] = 'success'
    except  Exception as e:
        response['msg'] = 'failure'
    return JsonResponse(response)

#教师选题方式2(根据科目名称筛选)
@require_http_methods(["GET"])
def show_questions_method2(request):
    response = {}
    book = xlrd.open_workbook('C:\\excel\\question_bank.xlsx')
    sheet1 = book.sheets()[0]
    nrows = sheet1.nrows #题库表格总行数
    Chinese_name = {'Math': '数学', 'Chinese': '语文', 'English': '英语'}
    single = []
    mul = []
    completion = []
    judgement = []
    subjective = []
    total = []
    try:
        questions = Question.objects.filter(subject = request.GET.get('subject'))
    except Exception as e:
        response['msg'] = 'question_failure' 
        return JsonResponse(response)
    for i in range(0, questions.count()):
        if(questions[i].types == "单选"):
            c = questions[i].location - 1
            d = {"location": c+1, "question": sheet1.cell(c,1).value, "choice1": sheet1.cell(c,2).value, 
            "choice2": sheet1.cell(c,3).value, "choice3": sheet1.cell(c,4).value, "choice4": sheet1.cell(c,5).value
            ,"answer": sheet1.cell(c,6).value, "description": sheet1.cell(c,7).value}
            single.append(d)
            total.append(d)
        elif(questions[i].types == "多选"):
            c = questions[i].location - 1
            d = {"location": c+1, "question": sheet1.cell(c,1).value, "choice1": sheet1.cell(c,2).value, 
            "choice2": sheet1.cell(c,3).value, "choice3": sheet1.cell(c,4).value, "choice4": sheet1.cell(c,5).value,
            "choice5": sheet1.cell(c,6).value, "answer": sheet1.cell(c,7).value, "description": sheet1.cell(c,8).value}
            mul.append(d)
            total.append(d)
        elif(questions[i].types == "填空"):
            c = questions[i].location - 1
            d = {"location": c+1, "question": sheet1.cell(c,1).value, "answer": sheet1.cell(c,2).value, 
            "description": sheet1.cell(c,3).value}
            completion.append(d)
            total.append(d)
        elif(questions[i].types == "判断"):
            c = questions[i].location - 1
            d = {"location": c+1, "question": sheet1.cell(c,1).value, "answer": sheet1.cell(c,2).value, 
            "description": sheet1.cell(c,3).value}
            judgement.append(d)
            total.append(d)
        elif(questions[i].types == "主观"):
            c = questions[i].location - 1
            d = {"location": c+1, "question": sheet1.cell(c,1).value, "description": sheet1.cell(c,2).value}
            subjective.append(d)
            total.append(d)
    response['single'] = single
    response['mul'] = mul
    response['completion'] = completion
    response['judgement'] = judgement
    response['subjective'] = subjective
    response['total'] = total
    response['msg'] = 'success'
    return JsonResponse(response)

#教师添题
@require_http_methods(["GET"])
def add_questions(request):
    response = {}
    Chinese_name = {'single': '单选', 'mul': '多选', 'completion': '填空', 'judgement': '判断', 'subjective': '主观', 'Math': '数学', 'Chinese': '语文', 'English': '英语'}
    #读取题库
    book = xlrd.open_workbook('C:\\excel\\question_bank.xlsx')
    sheet1 = book.sheets()[0]
    nrows = sheet1.nrows #题库表格总行数
    #写入题库
    wb = load_workbook('C:\\excel\\question_bank.xlsx')
    wb1 = wb.active #激活sheet
    try:
        types = request.GET.get('type')
        subject = request.GET.get('subject')
        difficulty = request.GET.get('difficulty')
        if(types == "single"): #单选
            question = request.GET.get('question') #题干
            choice1 = request.GET.get('choice1')
            choice2 = request.GET.get('choice2')
            choice3 = request.GET.get('choice3')
            choice4 = request.GET.get('choice4')
            answer = request.GET.get('answer')
            description = request.GET.get('description')
            try:
                que = Question(types = Chinese_name[types], subject = Chinese_name[subject], difficulty = difficulty, location = nrows + 1)
                que.save()
            except Exception as e:
                response['msg'] = 'insert_failure'
            try:
                wb1.cell(nrows+1, 1, nrows+1)
                wb1.cell(nrows+1, 2, question)
                wb1.cell(nrows+1, 3, choice1)
                wb1.cell(nrows+1, 4, choice2)
                wb1.cell(nrows+1, 5, choice3)
                wb1.cell(nrows+1, 6, choice4)
                wb1.cell(nrows+1, 7, answer)
                wb1.cell(nrows+1, 8, description)
                wb1.cell(nrows+1, 9, Chinese_name[types])
                wb.save('C:\\excel\\question_bank.xlsx') #保存
                response['msg'] = 'success'
            except Exception as e:
                response['msg'] = 'insert_failure2'
        elif(types == "mul"): #多选
            question = request.GET.get('question') #题干
            choice1 = request.GET.get('choice1')
            choice2 = request.GET.get('choice2')
            choice3 = request.GET.get('choice3')
            choice4 = request.GET.get('choice4')
            choice5 = request.GET.get('choice5')
            answer = request.GET.get('answer')
            description = request.GET.get('description')
            try:
                que = Question(types = Chinese_name[types], subject = Chinese_name[subject], difficulty = difficulty, location = nrows + 1)
                que.save()
            except Exception as e:
                response['msg'] = 'insert_failure'
            try:
                wb1.cell(nrows+1, 1, nrows+1)
                wb1.cell(nrows+1, 2, question)
                wb1.cell(nrows+1, 3, choice1)
                wb1.cell(nrows+1, 4, choice2)
                wb1.cell(nrows+1, 5, choice3)
                wb1.cell(nrows+1, 6, choice4)
                wb1.cell(nrows+1, 7, choice5)
                wb1.cell(nrows+1, 8, answer)
                wb1.cell(nrows+1, 9, description)
                wb1.cell(nrows+1, 10, Chinese_name[types])
                wb.save('C:\\excel\\question_bank.xlsx') #保存
                response['msg'] = 'success'
            except Exception as e:
                response['msg'] = 'insert_failure2'
        elif(types == "completion"): #填空
            question = request.GET.get('question') #题干
            answer = request.GET.get('answer')
            description = request.GET.get('description')
            try:
                que = Question(types = Chinese_name[types], subject = Chinese_name[subject], difficulty = difficulty, location = nrows + 1)
                que.save()
            except Exception as e:
                response['msg'] = 'insert_failure'
            try:
                wb1.cell(nrows+1, 1, nrows+1)
                wb1.cell(nrows+1, 2, question)
                wb1.cell(nrows+1, 3, answer)
                wb1.cell(nrows+1, 4, description)
                wb1.cell(nrows+1, 5, Chinese_name[types])
                wb.save('C:\\excel\\question_bank.xlsx') #保存
                response['msg'] = 'success'
            except Exception as e:
                response['msg'] = 'insert_failure2'
        elif(types == "judgement"): #判断
            question = request.GET.get('question') #题干
            answer = request.GET.get('answer')
            description = request.GET.get('description')
            try:
                que = Question(types = Chinese_name[types], subject = Chinese_name[subject], difficulty = difficulty, location = nrows + 1)
                que.save()
            except Exception as e:
                response['msg'] = 'insert_failure'
            try:
                wb1.cell(nrows+1, 1, nrows+1)
                wb1.cell(nrows+1, 2, question)
                wb1.cell(nrows+1, 3, answer)
                wb1.cell(nrows+1, 4, description)
                wb1.cell(nrows+1, 5, Chinese_name[types])
                wb.save('C:\\excel\\question_bank.xlsx') #保存
                response['msg'] = 'success'
            except Exception as e:
                response['msg'] = 'insert_failure2'
        elif(types == "subjective"): #主观
            question = request.GET.get('question') #题干
            description = request.GET.get('description')
            try:
                que = Question(types = Chinese_name[types], subject = Chinese_name[subject], difficulty = difficulty, location = nrows + 1)
                que.save()
            except Exception as e:
                response['msg1'] = 'insert_failure'
            try:
                wb1.cell(nrows+1, 1, nrows+1)
                wb1.cell(nrows+1, 2, question)
                wb1.cell(nrows+1, 3, description)
                wb1.cell(nrows+1, 4, Chinese_name[types])
                wb.save('C:\\excel\\question_bank.xlsx') #保存
                response['msg'] = 'success'
            except Exception as e:
                response['msg2'] = str(e)
    except Exception as e:
        response['msg'] = 'getType_failure'
    return JsonResponse(response)
    

#获取学生列表
@require_http_methods(["GET"])
def show_students(request):
    response = {}
    try:
        students = Student.objects.filter()   #没有设置过滤，即查询所有结果
        response['list'] = json.loads(serializers.serialize("json", students))
        response['msg'] = 'success'
    except  Exception as e:
        response['msg'] = 'failure'
    return JsonResponse(response)

#获取题库信息
@require_http_methods(["GET"])
def show_question_bank(request):
    response = {}
    username = request.GET.get('username')
    try:
        teacher = Teacher.objects.get(username =username)
    except  Exception as e:
        response['msg'] = 'teacher_failure' #没有该老师
        return JsonResponse(response)
    names = []
    question_banks = Question_bank.objects.filter(teacher = teacher)
    for question_bank in question_banks:
        names.append(question_bank.name)
    response['list'] = names
    return JsonResponse(response)

#获取题库详细信息
@require_http_methods(["GET"])
def show_question_bank_infor(request):
    response = {}
    try:
        teacher = Teacher.objects.get(username = request.GET.get('username'))
        question_bank = Question_bank.objects.get(name = request.GET.get('name'), teacher = teacher)
    except Exception as e:
        response['msg'] = str(e)
        return JsonResponse(response)
    try: #读取题库
        book = xlrd.open_workbook('C:\\excel\\question_bank.xlsx')
        sheet1 = book.sheets()[0]
        nrows = sheet1.nrows #题库表格总行数
    except Exception as e:
        response['msg'] = 'questionBank_failure' #读取题库遇到错误
        return JsonResponse(response)
    questions = question_bank.question.all() #该题库的所有题目
    single = []
    mul = []
    completion = []
    judgement = []
    subjective = []
    total = []
    num = questions.count()
    for i in range(0, num):
        if(questions[i].types == "单选"):
            c = questions[i].location - 1
            d = {"location": c+1, "question": sheet1.cell(c,1).value, "choice1": sheet1.cell(c,2).value, 
            "choice2": sheet1.cell(c,3).value, "choice3": sheet1.cell(c,4).value, "choice4": sheet1.cell(c,5).value
            ,"answer": sheet1.cell(c,6).value, "description": sheet1.cell(c,7).value}
            single.append(d)
            total.append(d)
        elif(questions[i].types == "多选"):
            c = questions[i].location - 1
            d = {"location": c+1, "question": sheet1.cell(c,1).value, "choice1": sheet1.cell(c,2).value, 
            "choice2": sheet1.cell(c,3).value, "choice3": sheet1.cell(c,4).value, "choice4": sheet1.cell(c,5).value,
            "choice5": sheet1.cell(c,6).value, "answer": sheet1.cell(c,7).value, "description": sheet1.cell(c,8).value}
            mul.append(d)
            total.append(d)
        elif(questions[i].types == "填空"):
            c = questions[i].location - 1
            d = {"location": c+1, "question": sheet1.cell(c,1).value, "answer": sheet1.cell(c,2).value, 
            "description": sheet1.cell(c,3).value}
            completion.append(d)
            total.append(d)
        elif(questions[i].types == "判断"):
            c = questions[i].location - 1
            d = {"location": c+1, "question": sheet1.cell(c,1).value, "answer": sheet1.cell(c,2).value, 
            "description": sheet1.cell(c,3).value}
            judgement.append(d)
            total.append(d)
        elif(questions[i].types == "主观"):
            c = questions[i].location - 1
            d = {"location": c+1, "question": sheet1.cell(c,1).value, "description": sheet1.cell(c,2).value}
            subjective.append(d)
            total.append(d)
    response['single'] = single
    response['mul'] = mul
    response['completion'] = completion
    response['judgement'] = judgement
    response['subjective'] = subjective
    response['total'] = total
    response['msg'] = 'success'
    return JsonResponse(response)

#创建题库
@require_http_methods(["GET"])
def create_question_bank(request):
    response = {}
    username = request.GET.get('username')
    try:
        teacher = Teacher.objects.get(username =username)
    except  Exception as e:
        response['msg'] = 'teacher_failure' #没有该老师
        return JsonResponse(response)
    try:
        question_bank = Question_bank(
            name = request.GET.get('name'),
            teacher = teacher
        )
        question_bank.save()
        response['msg'] = 'success'
    except  Exception as e:
        response['msg'] = 'failure'
    return JsonResponse(response)

#教师往自己的题库中插入题目
@require_http_methods(["POST"])
def insert_question_bank(request):
    response = {}
    result = json.loads(request.body.decode())
    try:
        question_bank = Question_bank.objects.get(name = result.get('name'))
    except Exception as e:
        response['msg'] = 'bankName_failure'
        return JsonResponse(response)
    locations = result.get('location') #获取到的题目包
    print(locations)
    for i in range(0, len(locations)):
        question = Question.objects.get(location = locations[i])
        question_bank.question.add(question)
    response['msg'] = 'success'
    return JsonResponse(response)

#教师创建题库(附带将选中的题目插入题库)
@require_http_methods(["POST"])
def create_question_bank_addQuestion(request):
    response = {}
    result = json.loads(request.body.decode())
    username = result.get('username')
    try:
        teacher = Teacher.objects.get(username = username)
    except  Exception as e:
        response['msg'] = 'teacher_failure' #没有该老师
        return JsonResponse(response)
    try:
        question_bank = Question_bank(
            name = result.get('name'),
            teacher = teacher
        )
        question_bank.save()
    except  Exception as e:
        response['msg'] = 'createBank_failure'
        return JsonResponse(response)
    try:
        question_bank = Question_bank.objects.get(name = result.get('name'))
    except Exception as e:
        response['msg'] = 'bankName_failure'
        return JsonResponse(response)
    locations = result.get('location') #获取到的题目包
    for i in range(0, len(locations)):
        question = Question.objects.get(location = locations[i])
        question_bank.question.add(question)
    response['msg'] = 'success'
    return JsonResponse(response)

#显示教师应该批改的主观题
@require_http_methods(["GET"])
def show_sub_homework(request):
    response={}
    try:
        teacher = Teacher.objects.get(username = request.GET.get('username'))
    except Exception as e:
        response['msg'] = 'teacher_failure'
        return JsonResponse(response)
    sub_answers = Sub_answer.objects.filter(homework__classes__teacher = teacher)
    question = []
    for sub_answer in sub_answers:
        d = {"homework": sub_answer.homework.name, "student": sub_answer.student.realname, "student_username": sub_answer.student.username}
        if(Homework_score.objects.filter(homework = sub_answer.homework, student = sub_answer.student).count() == 0): #没有该学生对应该题的分数信息
            question.append(d)
    response['homework'] = question 
    return JsonResponse(response)

#教师点击某个学生做的主观题进入批改界面
@require_http_methods(["GET"])
def show_sub_homework_infor(request):
    response={}
    book = xlrd.open_workbook('C:\\excel\\question_bank.xlsx')
    sheet1 = book.sheets()[0]
    nrows = sheet1.nrows #题库表格总行数
    try:
        homework = Homework.objects.get(name = request.GET.get('name'))
        student = Student.objects.get(username = request.GET.get('username'))
    except Exception as e:
        response['msg'] = 'getInformation_failure' #查找作业或学生失败
        return JsonResponse(response)
    sub_answer = Sub_answer.objects.get(student = student, homework = homework)
    c = sub_answer.question.location
    question = sheet1.cell(c-1,1).value 
    answer = sub_answer.answer 
    response['question'] = question #题干
    response['answer'] = answer #学生答案
    return JsonResponse(response)

#教师批改主观题
@require_http_methods(["GET"])
def correct_sub_homework(request):
    response={}
    try:
        homework = Homework.objects.get(name = request.GET.get('homework_name'))
    except Exception as e:
        response['msg'] = 'homework_failure'
        return JsonResponse(response)
    try:
        student = Student.objects.get(username = request.GET.get('student_username'))
    except Exception as e:
        response['msg'] = 'student_failure'
        return JsonResponse(response)
    try:
        homework_score = Homework_score(
            homework = homework,
            student = student,
            score = request.GET.get('score'),
            comment = request.GET.get('comment')
        )
        homework_score.save()
        response['msg'] = 'success'
    except Exception as e:
        response['msg'] = 'score_failure'
    return JsonResponse(response)

#教师查看历史作业
@require_http_methods(["GET"])
def show_homework_history(request):
    response={}
    history = []
    classes = Classes.objects.filter(name = request.GET.get('name'))
    for cl_ass in classes:
        homeworks = Homework.objects.filter(classes = cl_ass)
        for homework in homeworks:
            if(homework.close_time.replace(tzinfo=None).__lt__(datetime.datetime.now())):
                d = {"homework": homework.name, "create_time": homework.create_time.strftime("%Y-%m-%d %H:%M:%S"), 
                "close_time": homework.close_time.strftime("%Y-%m-%d %H:%M:%S"),"status":'已到期'}
            else:
                d = {"homework": homework.name, "create_time": homework.create_time.strftime("%Y-%m-%d %H:%M:%S"), 
                "close_time": homework.close_time.strftime("%Y-%m-%d %H:%M:%S"),"status":'未到期'}
            history.append(homework)
    response['history'] = history
    return JsonResponse(response)

#教师查看历史作业完成情况
@require_http_methods(["GET"])
def show_homework_situation(request):
    response = {}
    try:
        homework = Homework.objects.get(name = request.GET.get('homework_name'))
    except Exception as e:
        response['msg'] = 'homework_failure' #没有该作业
        return JsonResponse(response)
    
    score=[]
    homework_scores = Homework_score.objects.filter(homework = homework).order_by('-score')
    for homework_score in homework_scores:
        d = {"student": homework_score.student.realname,"score": homework_score.score,"time": homework_score.create_time}
        score.append(d)
    response['score'] = score
    return JsonResponse(response)